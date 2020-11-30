import openpyxl
from os import path

def load_workbook(wb_path):
    if path.exists(wb_path):
        return openpyxl.load_workbook(wb_path)
    return openpyxl.Workbook()

wb_path = 'Base.xlsx'
wb = load_workbook(wb_path)

sheet = wb['Planilha1']

#sheet.insert_rows(1,2)

#for col_idx, title in enumerate(('S.No.','Hobby'), start=1):
#    sheet.cell(row=1, column=col_idx + 1, value=title) - adicionar titulo na planilha (como se fosse filtro)


#sheet.delete_rows() - colocar numero da linha
#sheet.delete_cols() - colocar numero da coluna

#sheet.insert_cols(3)

#for row in sheet.values:
#    print(row)

linhas = sheet.capitalize()

wb.save(wb_path)

